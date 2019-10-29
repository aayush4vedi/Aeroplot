import openpyxl 
import random

db = openpyxl.load_workbook('db_copy.xlsx')
sheet = db.active

for r in range(2, sheet.max_row+1):
    for c in range(2,sheet.max_column+1):
        price = sheet.cell(row = r, column = c).value
        if(price == 'ERROR' or price is None):
            price = sheet.cell(row = r-1, column = c).value
            price = str(price)
            price = int(price.replace(',' , ''))
            price = price + random.randrange(100, 600)
            sheet.cell(row = r, column = c).value = price
        else:
            # print('else @row#{} col#{}, type: {}'.format(r,c,type(price)))
            price = str(price)
            price = int(price.replace(',' , ''))
            sheet.cell(row = r, column = c).value = price


db.save('db.xlsx')
print("======================Database has been cleaned=========================")