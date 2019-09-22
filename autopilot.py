import openpyxl 

text = open('flights.txt')
text.seek(0)

wb = openpyxl.load_workbook('flights.xlsx')
sheet = wb.active
sheet.delete_rows(0,sheet.max_row+1)
headers = ["ID","Date","Day","Day Type","Flight Number","From","To","Dep Time"]
sheet.append(headers)

for t in text.readlines()[2:]:
    t = t.split("|")
    entry = []
    for  e in t[1:]:
        entry.append(e.strip())
    sheet.append(entry)


wb.save('flights.xlsx')
print("======================Flight data has been updated======================")

#clense the old db

db = openpyxl.load_workbook('db.xlsx')
sheet = db.active
sheet.delete_rows(0,sheet.max_row+1)

#resetting the headers
sheet.cell(row = 1, column = 1).value = "TimeStamp"
for i in range(2,66):
        sheet.cell(row = 1, column = i).value = "f" + str(i-1)

db.save('db.xlsx')
print("======================Database has been restted=========================")