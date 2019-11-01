from datetime import datetime
import openpyxl 
import json

def datestr2int(datestring):
    datestring = str(datestring)
    d = datetime.strptime(datestring, '%Y-%m-%d %H:%M:%S.%f')
    s = datetime.strftime(d, '%m%d%H')
    return int(s)

data = {}
d = []
db = openpyxl.load_workbook('db.xlsx')
sheet = db.active

for c in range(2,sheet.max_column+1):
    datum = []
    f_id = sheet.cell(row = 1, column = c).value
    for r in range(2, sheet.max_row+1):
        entity = {}
        entity['date'] = datestr2int(sheet.cell(row = r, column = 1).value)
        entity['price'] = sheet.cell(row = r, column = c).value
        datum.append(entity)
    #data.append(datum)
    data[f_id] = datum
    d.append(data)

f_data = {}
f_data['flights'] = d

# print(data);
j = json.dumps(f_data, indent=4)
f = open('data.json', 'w')
# print(j, file=f)
print >> f, j
f.close()

print('=======================json uploaded=======================')