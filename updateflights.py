import openpyxl

# README.md to excel
data = """| ID  | Date       |  Day  |   Day Type      | Flight Number |  From  |   To   |Dep.Time| 
|---- |-----      | ---- |---------------  |--------------| :----: | :----: |------ |
| f1  | 27-09-2019 |  Fri  | Around Holiday  | Indigo-6E2625 |  Blr   |   Del  | 0025   |
| f2  | 27-09-2019 |  Fri  | Around Holiday  | GoAir-G8118   |  Blr   |   Del  | 2055   |
| f3  | 28-09-2019 |  Sat  | Around Holiday  | GoAir-G8118   |  Blr   |   Del  | 2055   |
| f4  | 28-09-2019 |  Sat  | Around Holiday  | SpiceJ-SG6638 |  Blr   |   Del  | 2310   |
| f5  | 3-10-2019  |  Tue  | Around Holiday  | GoAir-G8113   |  Del   |   Blr  | 0550   |
| f6  | 3-10-2019  |  Tue  | Around Holiday  | AirIn-AI933   |  Del   |   Blr  | 0510   |
| f7  | 4-10-2019  |  Fri  | Around Holiday  | Indigo-6E2808 |  Blr   |   Del  | 1930   |
| f8  | 4-10-2019  |  Fri  | Around Holiday  | SpiceJ-SG6638 |  Blr   |   Del  | 2310   |
| f9  | 7-10-2019  |  Mon  | Around Holiday  | SpiceJ-SG198  |  Blr   |   Del  | 2045   |
| f10 | 7-10-2019  |  Mon  | Around Holiday  | Indigo-6E2305 |  Blr   |   Del  | 2125   |
| f11 | 9-10-2019  |  Wed  | Around Holiday  | Indigo-6E2042 |  Del   |   Blr  | 0545   |
| f12 | 9-10-2019  |  Wed  | Around Holiday  | GoAir-G8113   |  Del   |   Blr  | 0550   |
| f13 | 24-10-2019 |  Fri  | Around Holiday  | AirIn-AI587   |  Blr   |   Del  | 2200   |
| f14 | 24-10-2019 |  Fri  | Around Holiday  | GoAir-G8118   |  Blr   |   Del  | 2055   |
| f15 | 25-10-2019 |  Fri  | Around Holiday  | GoAir-G8518   |  Blr   |   Del  | 2255   |
| f16 | 25-10-2019 |  Fri  | Around Holiday  | Indigo-6E2716 |  Blr   |   Del  | 2000   |
| f17 | 29-10-2019 |  Fri  | Around Holiday  | Indigo-6E5031 |  Del   |   Blr  | 0505   |
| f18 | 29-10-2019 |  Fri  | Around Holiday  | AirIn-AI933   |  Del   |   Blr  | 0510   |
| f19 | 30-10-2019 |  Fri  | Around Holiday  | Indigo-6E2701 |  Del   |   Blr  | 0655   |
| f20 | 30-10-2019 |  Fri  | Around Holiday  | GoAir-G8207   |  Del   |   Blr  | 0500   |
| f21 | 25-09-2019 |  Wed  |     Normal      | Indigo-6E5031 |  Del   |   Blr  | 0505   |
| f22 | 25-09-2019 |  Wed  |     Normal      | AirIn-AI851   |  Del   |   Blr  | 0455   |
| f23 | 25-09-2019 |  Wed  |     Normal      | GoAir-G8518   |  Blr   |   Del  | 2255   |
| f24 | 25-09-2019 |  Wed  |     Normal      | Indigo-6E2716 |  Blr   |   Del  | 2000   |
| f25 | 10-10-2019 |  Thu  |     Normal      | SpiceJ-SG8720 |  Blr   |   Del  | 2200   |
| f26 | 10-10-2019 |  Thu  |     Normal      | GoAir-G8118   |  Blr   |   Del  | 2055   |
| f27 | 10-10-2019 |  Thu  |     Normal      | Indigo-6E683  |  Del   |   Blr  | 0735   |
| f28 | 10-10-2019 |  Thu  |     Normal      | GoAir-G82610  |  Del   |   Blr  | 0810   |
| f29 | 7-11-2019  |  Thu  |     Normal      | Indigo-6E5031 |  Del   |   Blr  | 0505   |
| f30 | 7-11-2019  |  Thu  |     Normal      | AirIn-AI933   |  Del   |   Blr  | 0510   |
| f31 | 7-11-2019  |  Thu  |     Normal      | Indigo-6E2484 |  Blr   |   Del  | 2330   |
| f32 | 7-11-2019  |  Thu  |     Normal      | AirIn-AI610   |  Blr   |   Del  | 1910   | """

data = data.split("\n")
ids = []
dates = []
days = []
day_types = []
flight_numbers = []
froms =[]
tos = []
dep_times = []

wb = openpyxl.load_workbook('aeroplot.xlsx')
sheet = wb.active
sheet.cell(row = 1, column = 1).value = "ID"
sheet.cell(row = 1, column = 2).value = "Date"
sheet.cell(row = 1, column = 3).value = "Day"
sheet.cell(row = 1, column = 4).value = "Day Type"
sheet.cell(row = 1, column = 5).value = "Flight Number"
sheet.cell(row = 1, column = 6).value = "From"
sheet.cell(row = 1, column = 7).value = "To"
sheet.cell(row = 1, column = 8).value = "Dep Time"

data = data[2:]
r = 2
for d in data:
    d = d.split("|")
    ids.append(d[1].strip())
    sheet.cell(row = r, column = 1).value = d[1].strip()
    days.append(d[2].strip())
    sheet.cell(row = r, column = 2).value = d[2].strip()
    dates.append(d[3].strip())
    sheet.cell(row = r, column = 3).value = d[3].strip()
    day_types.append(d[4].strip())
    sheet.cell(row = r, column = 4).value = d[4].strip()
    flight_numbers.append(d[5].strip())
    sheet.cell(row = r, column = 5).value = d[5].strip()
    froms.append(d[6].strip())
    sheet.cell(row = r, column = 6).value = d[6].strip()
    tos.append(d[7].strip())
    sheet.cell(row = r, column = 7).value = d[7].strip()
    dep_times.append(d[8].strip())
    sheet.cell(row = r, column = 8).value = d[8].strip()
    r += 1

wb.save('aeroplot.xlsx')